import openpyxl #the needed imports for this script
from sys import argv

script, filename = argv #sets the script name and filename as necessary args to run
                        #IMPROVEMENT: set the filename as a rawinput to pass

data = open(filename, "r") #gives a variable for the opened file

everything = []
title = []
expense = []                #***The necessary lists we'll be using***
reimbursable = []           # Not sure how necessary or good it is to just throw these
living_expense = []         # all out there at once, but it helped me wrap my head around
housing = []                # what I was doing
other = []

def sort_data(f): #first function, takes our open file, sorts its data between a price (integer), and its title. Into two lists
    all_info = f.read().split()
    while len(all_info) > 0:
        an_item = all_info.pop()
        try:
            expense.append(int(an_item))
        except ValueError:
            title.append(an_item)

sort_data(data) #runs the sort_data function on our open file with our budjet data

def process_lists(a, b): # this function takes our new expense and title lists, prints
                            # corresonding values side by side, and allows me to input
    position = 0            # what category of expense(and its list) I want to send it to(ie housing costs)


    while len(a) and len(b) > 0:
        item_title = a[position]
        item_expense = b[position]

        print "Where would you like to send this one ?"
        print """******************"""
        print item_title,
        print item_expense
        print """******************"""
        send = raw_input("A = Reimburse, B = Living Expense, C = Housing, D = Other  ")

        if send == "A":
            title.remove(item_title) # removes items from their master expense/title list
            expense.remove(item_expense)
            reimbursable.append(item_title), #appends both items to their appropriate list based
            reimbursable.append(item_expense)#on what excel sheet it will be recorded on

            while item_title in title:          #this loop will double check expenses with same
                locT = title.index(item_title)  #..title and append them too, this was to avoid
                locE = title.index(item_title)  #..processing redundant expenses that
                popTitle = title.pop(locT)      #..were tracked a lot, like transportation and coffee
                popExpense = expense.pop(locE)
                reimbursable.append(popTitle)
                reimbursable.append(popExpense)

        elif send == "B":
            title.remove(item_title)
            expense.remove(item_expense)
            living_expense.append(item_title),
            living_expense.append(item_expense)

            while item_title in title:
                locT = title.index(item_title) # had to find a way to keep the expense
                locE = title.index(item_title) # ..and title aligned and appending
                popTitle = title.pop(locT)     #..together since they were techically in two
                popExpense = expense.pop(locE) # seperate lists.
                living_expense.append(popTitle)
                living_expense.append(popExpense)


        elif send == "C":
            title.remove(item_title)
            expense.remove(item_expense)
            housing.append(item_title),
            housing.append(item_expense)

            while item_title in title:
                locT = title.index(item_title)
                locE = title.index(item_title)
                popTitle = title.pop(locT)
                popExpense = expense.pop(locE)
                housing.append(popTitle)
                housing.append(popExpense)

        elif send == "D":
            title.remove(item_title)
            expense.remove(item_expense)
            other.append(item_title),
            other.append(item_expense)

            while item_title in title:
                locT = title.index(item_title)
                locE = title.index(item_title)
                popTitle = title.pop(locT)
                popExpense = expense.pop(locE)
                other.append(popTitle)
                other.append(popExpense)

        else:
            print "Unrecognized input, try again."

process_lists(title, expense)

budj = openpyxl.load_workbook('BlankBudj.xlsx')  #sets variables for the Excel workbook I'll be using (blank template already saved in directory)
                                                #and the following names and gives variables for Sheets, there will be a sheet for each expense type
sheet_main = budj.active                           # this is to cross check specific expenses if I want to
sheet_reim = budj.get_sheet_by_name('Sheet2')
sheet_living = budj.get_sheet_by_name('Sheet3')
sheet_housing = budj.create_sheet(title = 'Housing')
sheet_other = budj.create_sheet(title = 'Other')

sheet_reim.title = 'Reimbursements'
sheet_living.title = 'Living'
sheet_housing.title = 'Housing'
sheet_main.title = 'Main'

def list_to_sheet(a_list, a_sheet): #this function will cycle through our sorted lists, and append them to the proper expense sheet in workbook
    rows_needed = len(a_list) / 2 + 1

    while len(a_list) > 0:
        for rowNum in range(1, rows_needed):
            for colNum in range(1, 3):
                if len(a_list) > 0:
                    our_thing = a_list.pop()
                    a_sheet.cell(row = rowNum, column = colNum).value = our_thing


list_to_sheet(reimbursable, sheet_reim)
list_to_sheet(living_expense, sheet_living)
list_to_sheet(housing, sheet_housing)
list_to_sheet(other, sheet_other) #runs the function for every list, sheet combo we have

def total_sheets(from_sheet): #this function will take values from column 1 of a expense sheet, and get their sum
    rows = from_sheet.max_row
    to_total =[]
    for x in range(1, rows):
        a_value = from_sheet.cell(row = x, column = 1).value
        to_total.append(a_value)

    return sum(to_total)

print "What is the conversion rate? "
conv_rate = raw_input("? ")  #the conversion rate here is changing daily, this helps me adjust when Im months behind
sheet_main['C2'].value = total_sheets(sheet_living) / int(conv_rate) #takes the cell in the main sheet, will make its value the corresponding sum of the numbers in the expense sheet that goes with it
sheet_main['C12'].value = total_sheets(sheet_housing) / int(conv_rate)
sheet_main['C20'].value = total_sheets(sheet_other) / int(conv_rate)
sheet_main['C35'].value = total_sheets(sheet_reim) / int(conv_rate)


print "Save your new budjet sheet as? "
save_as = raw_input("> ")
budj.save(save_as)


#Ways to improve: a back function, a once sorted sort all of same value,
